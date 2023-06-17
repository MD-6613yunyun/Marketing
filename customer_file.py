# This script is importing into google contacts with the customer information excel workbook
from openpyxl import load_workbook
import csv
# Load the Excel file
excel_file = "C:\\Users\\Hp\\OneDrive\\Documents\\KoPyaeContacts\\MDY_Shop_Customer_List.xlsx"
csv_file = "C:\\Users\\Hp\\OneDrive\\Documents\\KoPyaeContacts\\contacts_two.csv"
workbook = load_workbook(filename=excel_file)
# Select the worksheet
worksheet = workbook['one']  # Replace 'Sheet1' with the actual sheet name

header = ['Name','Group Membership','Address 1 - Street','Phone 1 - Type','Phone 1 - Value','Address 1 - City']
# Read rows one by one
# counter = 0
# for row in worksheet.iter_rows(values_only=True):
#     # Process each row
#     if counter == 0:
#         row_data = list(row)[1:]
#         row_data[1] = row_data[1] + ' ::: * myContacts'
#         row_data[3] = str(row_data[3]).replace('9','+959',1)
#         row_data.insert(3,'Mobile')
#         print(row_data[:-1])
#     counter += 1


with open(csv_file, 'w', newline='',encoding='utf-8') as file:
    writer = csv.writer(file)
    
    # Write the header row
    writer.writerow(header)
    
    # Write the data rows
    counter = 0
    for row in worksheet.iter_rows(values_only=True):
        # Process each row
        if counter != 0:
            row_data = list(row)[:3]
            ph = str(row_data[2])
            if ph.startswith("09"):
                if "," in ph:
                    ph = ph.split(",")[0]
                elif "/" in ph:
                    ph = ph.split("/")[0]
                ph = ph.replace("09-","+959",1)
            else:
                ph = None
            row_data.insert(1,'Mdy_Customer ::: * myContacts')
            row_data.insert(3,'Mobile')
            row_data[4] = ph
            row_data.insert(5,'Mandalay')
            print(row_data)
            writer.writerow(row_data)
        counter += 1


